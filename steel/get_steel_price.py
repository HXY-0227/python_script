import requests
import re
import time
import datetime
import os
from lxml import etree
from openpyxl import Workbook
from openpyxl import load_workbook


class DailyMarketPage:
    """每日某材料的行情页面

    Attributes:
        current_date: 当前日期 
        url: url
        title: 标题
    """

    # 日期
    current_date = ''
    # 当前行情URL
    url = ''
    # 描述
    title = ''

class DailyMarketDetail:
    """某一天某材料的价格详情

    Attributes:
        data: 日期 类似1日 2日 对其Excel模板的sheet名称
        pipe_name: 材料名称
        specifications: 规格
        price: 当日价格
    """
    date = ''
    pipe_name = ''
    specifications = ''
    price = 0


header = ''
# cookie
cookies = {}
# 工作簿
wb = None
# 文件路径
file_path = ''
# 需要保存的sheet集合，模板一共有31天的，不是所有都需要留在最终生成的文件中
save_sheet = []
# 更多页面的URL集合
urls = []

def init():
    print('load excel template...')
    global header
    header = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36",
        "Referer": "http://list1.mysteel.com/market/p-236-----010109-0-01030204-------1.html"
    }
    cookie_str = ''

    global cookies
    cookies = {}
    for line in cookie_str.split(';'):
        key, value = line.split('=', 1)
        cookies[key] = value

    global urls
    # 焊管、无缝管、螺旋管、镀锌管行情列表 查询1-3页
    urls.append('http://list1.mysteel.com/market/p-236-----010109-0-01030204-------1.html')
    urls.append('http://list1.mysteel.com/market/p-236-----010109-0-01030204-------2.html')
    urls.append('http://list1.mysteel.com/market/p-236-----010109-0-01030204-------3.html')
    # 镀锌板卷列表 查询1-3页
    urls.append('https://list1.mysteel.com/market/p-435-0----010105-0-01030204-------1.html')
    urls.append('https://list1.mysteel.com/market/p-435-0----010105-0-01030204-------2.html')
    urls.append('https://list1.mysteel.com/market/p-435-0----010105-0-01030204-------3.html')
    # 热轧板卷列表 查询1-3页
    urls.append('https://list1.mysteel.com/market/p-231-----010103-0-01030204-------1.html')
    urls.append('https://list1.mysteel.com/market/p-231-----010103-0-01030204-------2.html')
    urls.append('https://list1.mysteel.com/market/p-231-----010103-0-01030204-------3.html')
    # 花纹板卷列表 查询1-3页
    urls.append('https://list1.mysteel.com/market/p-231-----010103-0-01010601-------1.html')
    urls.append('https://list1.mysteel.com/market/p-231-----010103-0-01010601-------2.html')
    urls.append('https://list1.mysteel.com/market/p-231-----010103-0-01010601-------3.html')
    # H型钢、工角槽钢列表 查询1-3页
    urls.append('https://list1.mysteel.com/market/p-223-----010107-0-01030204-------1.html')
    urls.append('https://list1.mysteel.com/market/p-223-----010107-0-01030204-------2.html')
    urls.append('https://list1.mysteel.com/market/p-223-----010107-0-01030204-------3.html')
    # 圆钢列表 查询1-3页
    urls.append('https://list1.mysteel.com/market/p-228-15346-----0--------1.html')
    urls.append('https://list1.mysteel.com/market/p-228-15346-----0--------2.html')
    urls.append('https://list1.mysteel.com/market/p-228-15346-----0--------3.html')
    # 方钢列表 查询1-3页
    urls.append('https://list1.mysteel.com/market/p-223-----010107-0-010101-------1.html')
    urls.append('https://list1.mysteel.com/market/p-223-----010107-0-010101-------2.html')
    urls.append('https://list1.mysteel.com/market/p-223-----010107-0-010101-------3.html')

    global file_path
    file_path = os.getcwd()

    global wb
    wb = None
    wb = load_workbook(filename=file_path + '\\武汉行政区域直管项目钢材信息价（和合盈升）模板.xlsx')
    print('load excel template successfully...')

def save_pipe(begin_time, end_time):
    """保存管材的价格信息

    Args:
        begin_time: 开始时间
        end_time: 结束时间
    """
    print('begin write all pipe price to excel, please wait...')
    all_page = get_all_pipe_data_url(begin_time, end_time)

    welded_pipe_data = []
    for daily_page in all_page['welded_pipe_market_list']:
        welded_pipe_data.append(get_daily_pipe_data(daily_page, '焊管', '4寸*3.75mm', './td[2]','./td[6]'))
    writeExcel(welded_pipe_data, 'N4')

    seamless_tube_data = []
    for daily_page in all_page['seamless_tube_market_list']:
        seamless_tube_data.append(get_daily_pipe_data(daily_page, '流体管', 'Ф159*6', './td[3]','./td[6]'))
    writeExcel(seamless_tube_data, 'N35')

    spiral_tube_data = []
    for daily_page in all_page['spiral_tube_market_list']:
        spiral_tube_data.append(get_daily_pipe_data(daily_page, '螺旋管', '426*6-8', './td[2]','./td[5]'))
    writeExcel(spiral_tube_data, 'N106')

    galvanized_pipe_data = []
    for daily_page in all_page['galvanized_pipe_market_list']:
        galvanized_pipe_data.append(get_daily_pipe_data(daily_page, '镀锌管', '4寸*3.75mm', './td[2]','./td[6]'))
    writeExcel(galvanized_pipe_data, 'N162')

    galvanized_coil_data = []
    for daily_page in all_page['galvanized_coil_list']:
        galvanized_coil_data.append(get_daily_pipe_data(daily_page, '镀锌板卷', '1.5*1250*C', './td[2]','./td[5]'))
    writeExcel(galvanized_coil_data, 'N204')

    hot_rolled_coil_data = []
    for daily_page in all_page['hot_rolled_coil_list']:
        hot_rolled_coil_data.append(get_daily_pipe_data(daily_page, '热轧板卷', '5.75-11.5*1500*C', './td[2]','./td[5]'))
    writeExcel(hot_rolled_coil_data, 'N225')

    pattern_coil_data = []
    for daily_page in all_page['pattern_coil_list']:
        pattern_coil_data.append(get_daily_pipe_data(daily_page, '花纹板卷', '3.5*1250*C', './td[2]','./td[5]'))
    writeExcel(pattern_coil_data, 'N259')

    h_beam_data = []
    for daily_page in all_page['h_beam_list']:
        h_beam_data.append(get_daily_pipe_data(daily_page, 'H型钢', '250*250*9*14', './td[2]','./td[5]'))
    writeExcel(h_beam_data, 'N269')

    angle_channel_steel_data = []
    for daily_page in all_page['angle_channel_steel_list']:
        angle_channel_steel_data.append(get_daily_pipe_data(daily_page, '工字钢', '20#', './td[2]','./td[5]'))
    writeExcel(angle_channel_steel_data, 'N297')
    
    round_bar_data = []
    for daily_page in all_page['round_bar_list']:
        round_bar_data.append(get_daily_pipe_data(daily_page, '圆钢', 'Ф16-25', './td[2]','./td[5]'))
    writeExcel(round_bar_data, 'N275')

    square_steel_data = []
    for daily_page in all_page['square_steel_list']:
        square_steel_data.append(get_daily_pipe_data(daily_page, '方钢', '20*20', './td[2]','./td[5]'))
    writeExcel(square_steel_data, 'N378')

    print('write all pipe price to excel successfully...')

def get_all_pipe_data_url(begin_time, end_time):
    """抓取武汉要抓取材料对应时间区间的URL
    
    Args:
        begin_time: 开始时间
        end_time: 结束时间
    
    Returns:
        字典 example key: welded_pipe_market_list  value: 焊管从开始时间到结束时间区间内每一天的价格行情URL
    """
    # 焊管列表
    welded_pipe_market_list = []
    # 无缝管列表
    seamless_tube_market_list = []
    # 螺旋管列表
    spiral_tube_market_list = []
    # 镀锌管列表
    galvanized_pipe_market_list = []
    # 镀锌板卷列表
    galvanized_coil_list = []
    # 热轧板卷
    hot_rolled_coil_list = []
    # 花纹板卷
    pattern_coil_list = []
    # H型钢
    h_beam_list = []
    # 工角槽钢
    angle_channel_steel_list = []
    # 圆钢
    round_bar_list = []
    # 方钢
    square_steel_list = []

    # 抓取数据
    for url in urls:
        response = requests.get(url, headers=header, cookies=cookies)
        response.encoding = 'gbk'
        wuhan_steel_pipe_market = response.text
        item_urls = re.findall('<a href="(.*?)" title=".*?" target=".*?">', wuhan_steel_pipe_market)
        titles = re.findall('<a href=".*?" title="(.*?)" target=".*?">', wuhan_steel_pipe_market)
        dates = re.findall('<span class="date">(.*?)</span>', wuhan_steel_pipe_market)
    
        for i in range(len(item_urls)):
            if (time.strftime(dates[i][0:10]).__ge__(time.strftime(begin_time)) 
                and time.strftime(dates[i][0:10]).__le__(time.strftime(end_time))):

                dailyMarketPage = DailyMarketPage()
                dailyMarketPage.current_date = dates[i][0:10]
                dailyMarketPage.title = titles[i]
                dailyMarketPage.url = item_urls[i]

                if '（' in dailyMarketPage.title:
                    continue

                global save_sheet
                save_sheet.append(dailyMarketPage.title.split('武')[0])

                if '焊管' in dailyMarketPage.title:
                    welded_pipe_market_list.append(dailyMarketPage)
                elif '无缝管' in dailyMarketPage.title:
                    seamless_tube_market_list.append(dailyMarketPage)
                elif '螺旋管' in dailyMarketPage.title:
                    spiral_tube_market_list.append(dailyMarketPage)
                elif '镀锌管' in dailyMarketPage.title:
                    galvanized_pipe_market_list.append(dailyMarketPage)
                elif '镀锌板卷' in dailyMarketPage.title:
                    galvanized_coil_list.append(dailyMarketPage)
                elif '热轧板卷' in dailyMarketPage.title:
                    hot_rolled_coil_list.append(dailyMarketPage)
                elif '花纹板卷' in dailyMarketPage.title:
                    pattern_coil_list.append(dailyMarketPage)
                elif 'H型钢' in dailyMarketPage.title:
                    h_beam_list.append(dailyMarketPage)
                elif '工角槽钢' in dailyMarketPage.title:
                    angle_channel_steel_list.append(dailyMarketPage)
                elif '建筑钢材' in dailyMarketPage.title:
                    round_bar_list.append(dailyMarketPage)
                elif '方钢' in dailyMarketPage.title:
                    square_steel_list.append(dailyMarketPage)

    pipe_url = {}
    pipe_url['welded_pipe_market_list'] = welded_pipe_market_list
    pipe_url['seamless_tube_market_list'] = seamless_tube_market_list
    pipe_url['spiral_tube_market_list'] = spiral_tube_market_list
    pipe_url['galvanized_pipe_market_list'] = galvanized_pipe_market_list
    pipe_url['galvanized_coil_list'] = galvanized_coil_list
    pipe_url['hot_rolled_coil_list'] = hot_rolled_coil_list
    pipe_url['pattern_coil_list'] = pattern_coil_list
    pipe_url['h_beam_list'] = h_beam_list
    pipe_url['angle_channel_steel_list'] = angle_channel_steel_list
    pipe_url['round_bar_list'] = round_bar_list
    pipe_url['square_steel_list'] = square_steel_list
    return pipe_url

def get_daily_pipe_data(dailyData, name, specifications, specifications_index, price_index):
    """通过某一天的URL，获取每一天某材料的平均价格

    Args:
        dailyData: 某日材料对应URL等信息
        name: 材料名称
        specifications: 规格
        specifications_index: 规格索引
        price_index: 价格索引，不同产品价格索引列不同

    Returns:
        某一天某材料的平均价格
    """

    # 抓取数据
    response = requests.get('http:' + dailyData.url, headers=header, cookies=cookies)
    dom_tree = etree.HTML(response.content)

    # 筛选对应规格的产品的所有价格
    daily_pipe_all_price = []
    for text in dom_tree.xpath("//table[@id='marketTable']/tr")[3:]:
        if ((text.xpath(specifications_index)[0].text.strip() == specifications)
            and (text.xpath("./td[1]")[0].text.strip() == name)):
            dailyMarketDetail = DailyMarketDetail()
            dailyMarketDetail.pipe_name = text.xpath("./td[1]")[0].text.strip()
            dailyMarketDetail.specifications = text.xpath(specifications_index)[0].text.strip()
            dailyMarketDetail.price = int(text.xpath(price_index)[0].text.strip())
            if dailyMarketDetail.pipe_name == '镀锌板卷' and '有花' in text.xpath('./td[7]')[0].text.strip():
                continue
            daily_pipe_all_price.append(dailyMarketDetail)
    
    # 对价格求平均值
    if len(daily_pipe_all_price) > 0:
        dailyMarketDetail = DailyMarketDetail()
        dailyMarketDetail.pipe_name = daily_pipe_all_price[0].pipe_name
        dailyMarketDetail.specifications = daily_pipe_all_price[0].specifications
        dailyMarketDetail.price = sum(item.price for item in daily_pipe_all_price) / len(daily_pipe_all_price)
        if dailyMarketDetail.pipe_name == '花纹板卷':
            dailyMarketDetail.date = dailyData.title.split('南')[0]
        elif dailyMarketDetail.pipe_name == '方钢':
            dailyMarketDetail.date = dailyData.title.split('上')[0]
        else:
            dailyMarketDetail.date = dailyData.title.split('武')[0]
        return dailyMarketDetail

def writeExcel(data, position):
    """写Excel

    Args:
        data: 要写入的数据集
        position: 单元格位置
    """
    for dailyData in data:
        if (dailyData != None):
            try:
                ws = wb[dailyData.date]
                ws[position] = dailyData.price
            except KeyError:
                print('filter the repeat data, ingore the message...')
        

if __name__ == "__main__":
    init()

    begin_time = input("please input begin time: ")
    end_time = input("please input end time: ")
    
    save_pipe(begin_time, end_time)

    all_sheet = wb.sheetnames
    for sheet in all_sheet:
        if sheet not in set(save_sheet):
            del wb[sheet]

    wb.save(filename=file_path + '\\武汉行政区域直管项目钢材信息价（和合盈升）.xlsx')
    print('save excel successfully...')