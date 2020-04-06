# -*- coding: utf-8 -*-
# @Time : 2020-04-06
# @Author : HXY
# @File : auto_excel.py

from openpyxl import load_workbook

# ID所在的列
DEFAULT_ID_COLUMN_NUM = 1

def count_ids():
    """
    统计骑手ID，并去重，拷贝到第二个表格
    """

    # 获取原始数据表中的的ids并去重
    workbook = load_workbook(filename='data.xlsx')
    sheet = workbook.active
    ids = []
    for cell in sheet['A']:
        if((cell.value is not None) and (cell.value not in ids)):
            ids.append(cell.value)

    # 创建统计结果sheet，写入id列
    workbook.create_sheet('统计结果1')
    count_result_sheet = workbook['统计结果1']
    index = 0
    for id in ids:
        index += 1
        count_result_sheet.cell(index, DEFAULT_ID_COLUMN_NUM).value = id

    # 填充姓名
    count_result_sheet['B1'] = '姓名'
    for index in range(2, len(ids) + 1):
        column_id = 'A' + str(index)
        column_name = 'B'+ str(index)
        count_result_sheet[column_name] = '=VLOOKUP(' + column_id + ',表格数据!A:B,2,0)'
       
        
    # 统计当天完成订单总和
    count_result_sheet['C1'] = '当天外卖完成单'
    for index in range(2, len(ids) + 1):
        column = 'C' + str(index)
        count_result_sheet[column] = '=SUMIFS(表格数据!D:D,表格数据!B:B,表格数据!B' + str(index) + ')'
    
    # 填充骑手等级
    count_result_sheet['D1'] = '骑手等级'
    for index in range(2, len(ids) + 1):
        column_id = 'A' + str(index)
        column_level = 'D' + str(index)
        count_result_sheet[column_level] = '=VLOOKUP(' + column_id + ',表格数据!A:F,5,0)'

    # 填充骑手分层
    count_result_sheet['E1'] = '骑手分层'
    for index in range(2, len(ids) + 1):
        column_id = 'A' + str(index)
        column_layer = 'E' + str(index)
        count_result_sheet[column_layer] = '=vLOOKUP(' + column_id + ',表格数据!A:F,6,0)'

    # 填充骑手等级
    count_result_sheet['F1'] = '骑手等级'
    for index in range(2, len(ids) + 1):
        column_origin = 'D' + str(index)
        column_level = 'F' + str(index)
        count_result_sheet[column_level] = '=MID(' + column_origin + ',3,2)'

    # 统计骑手等级，不知道为啥跨sheet，countif函数不好用
    # 统计骑手等级
    count_result_sheet['G1'] = '骑手等级'
    count_result_sheet['G2'] = '青铜'
    count_result_sheet['G3'] = '白银'
    count_result_sheet['G4'] = '黄金'
    count_result_sheet['G5'] = '王者'
    count_result_sheet['H1'] = '合计'
    count_result_sheet['H2'] = '=COUNTIF(F:F,G2)'
    count_result_sheet['H3'] = '=COUNTIF(F:F,G3)'
    count_result_sheet['H4'] = '=COUNTIF(F:F,G4)'
    count_result_sheet['H5'] = '=COUNTIF(F:F,G5)'

    for col in ['F', 'G', 'H']:
        count_result_sheet.column_dimensions[col].hidden = True
    
    # 保存文件
    workbook.save(filename='data.xlsx')


def count_other():
    """
    统计骑手等级和分层
    """

    workbook = load_workbook(filename='data.xlsx')
    workbook.create_sheet('统计结果2')
    count_rider_sheet = workbook['统计结果2']

    # 统计骑手等级
    count_rider_sheet['A1'] = '骑手等级'
    count_rider_sheet['A2'] = '青铜'
    count_rider_sheet['A3'] = '白银'
    count_rider_sheet['A4'] = '黄金'
    count_rider_sheet['A5'] = '王者'
    count_rider_sheet['B1'] = '合计'
    count_rider_sheet['B2'] = '=统计结果1!H2'
    count_rider_sheet['B3'] = '=统计结果1!H3'
    count_rider_sheet['B4'] = '=统计结果1!H4'
    count_rider_sheet['B5'] = '=统计结果1!H5'

    # 统计骑手分层
    count_rider_sheet['C1'] = '骑手分层（众包）'
    count_rider_sheet['C2'] = '全能骑手'
    count_rider_sheet['C3'] = '效能短板骑手'
    count_rider_sheet['C4'] = '出勤短板骑手'
    count_rider_sheet['C5'] = '时长短板骑手'
    count_rider_sheet['C6'] = '尾部骑手'
    count_rider_sheet['D1'] = '合计'
    count_rider_sheet['D2'] = '=COUNTIF(表格数据!F:F,C2)'
    count_rider_sheet['D3'] = '=COUNTIF(表格数据!F:F,C3)'
    count_rider_sheet['D4'] = '=COUNTIF(表格数据!F:F,C4)'
    count_rider_sheet['D5'] = '=COUNTIF(表格数据!F:F,C5)'
    count_rider_sheet['D6'] = '=COUNTIF(表格数据!F:F,C6)'

    workbook.save(filename='data.xlsx')

count_ids()
count_other()